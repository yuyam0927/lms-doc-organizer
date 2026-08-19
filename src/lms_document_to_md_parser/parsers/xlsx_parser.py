from pathlib import Path

from openpyxl import load_workbook

from ._markdown import rows_to_table


def _cell_to_str(value, formula) -> str:
    if value is not None:
        return str(value)
    if isinstance(formula, str) and formula.startswith("="):
        # No cached value (e.g. the workbook was never opened in a
        # spreadsheet app) -- show the formula rather than silently
        # dropping the cell's content.
        return formula
    return ""


def _row_to_cells(value_row, formula_row) -> list[str]:
    return [_cell_to_str(value, formula) for value, formula in zip(value_row, formula_row)]


def xlsx_to_markdown(path: Path) -> str:
    values_wb = load_workbook(str(path), data_only=True, read_only=True)
    formulas_wb = load_workbook(str(path), data_only=False, read_only=True)
    parts: list[str] = []

    try:
        for sheet_name in values_wb.sheetnames:
            value_sheet = values_wb[sheet_name]
            formula_sheet = formulas_wb[sheet_name]
            rows = [
                _row_to_cells(value_row, formula_row)
                for value_row, formula_row in zip(
                    value_sheet.iter_rows(values_only=True),
                    formula_sheet.iter_rows(values_only=True),
                )
            ]
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if not rows:
                continue

            width = max(len(row) for row in rows)
            rows = [row + [""] * (width - len(row)) for row in rows]

            table = rows_to_table(rows)
            parts.append(f"## {sheet_name}\n\n{table}")
    finally:
        values_wb.close()
        formulas_wb.close()

    return "\n\n".join(parts) + "\n"
