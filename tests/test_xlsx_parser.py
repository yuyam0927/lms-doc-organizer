from pathlib import Path

from openpyxl import Workbook

from lms_document_to_md_parser.parsers.xlsx_parser import xlsx_to_markdown


def _make_xlsx(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1.append(["h1", "h2"])
    ws1.append(["x|y", None])
    ws1.append([None, None])  # fully blank row should be dropped

    ws2 = wb.create_sheet("Empty")
    # ws2 has no rows appended -> should be skipped entirely

    wb.save(str(path))


def test_xlsx_to_markdown_renders_sheet_and_escapes_pipe(tmp_path):
    xlsx_path = tmp_path / "sample.xlsx"
    _make_xlsx(xlsx_path)

    md = xlsx_to_markdown(xlsx_path)

    assert "## Data" in md
    assert "| h1 | h2 |" in md
    assert "| x\\|y |  |" in md
    assert "## Empty" not in md


def test_xlsx_to_markdown_pads_short_rows(tmp_path):
    xlsx_path = tmp_path / "ragged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["a", "b", "c"])
    ws.append(["only-one"])
    wb.save(str(xlsx_path))

    md = xlsx_to_markdown(xlsx_path)

    lines = [line for line in md.splitlines() if line.startswith("|")]
    assert all(line.count("|") == 4 for line in lines)
